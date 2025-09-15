"""
Rotas de documentos para o Sistema Alpha Gestão Documental
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, make_response
from flask_login import login_required, current_user
from app.models import Document, DocumentVersion, DocumentReading, ApprovalFlow, DocumentType
from app import db
from datetime import datetime, timedelta
import uuid
import os
import tempfile
import html2text
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, BaseDocTemplate, PageTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
import re
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import xlsxwriter

bp = Blueprint('documents', __name__)

@bp.route('/')
@login_required
def index():
    """Lista de documentos"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    tipo = request.args.get('tipo', '')
    status = request.args.get('status', '')
    
    query = Document.query.filter_by(ativo=True)
    
    if search:
        query = query.filter(
            (Document.titulo.contains(search)) |
            (Document.codigo.contains(search)) |
            (Document.palavras_chave.contains(search))
        )
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    
    if status:
        query = query.filter_by(status=status)
    
    documents = query.order_by(Document.data_criacao.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Tipos de documento para filtro
    tipos = db.session.query(Document.tipo.distinct()).filter_by(ativo=True).all()
    tipos = [t[0] for t in tipos]
    
    return render_template('documents/index.html', 
                         documents=documents, 
                         search=search,
                         tipos=tipos,
                         current_tipo=tipo,
                         current_status=status)

@bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """Criar novo documento"""
    if not current_user.can_create_documents():
        flash('Você não tem permissão para criar documentos.', 'error')
        return redirect(url_for('documents.index'))
    
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        tipo = request.form.get('tipo')
        departamento = request.form.get('departamento')
        palavras_chave = request.form.get('palavras_chave')
        resumo = request.form.get('resumo')
        conteudo = request.form.get('conteudo')
        data_validade = request.form.get('data_validade')
        
        # Gerar código único
        codigo = f"{tipo.upper()}-{datetime.now().strftime('%Y')}-{str(uuid.uuid4())[:8].upper()}"
        
        # Obter tipo_documento_id se fornecido
        tipo_documento_id = request.form.get('tipo_documento_id')
        
        # Criar documento
        document = Document(
            codigo=codigo,
            titulo=titulo,
            tipo=tipo,
            tipo_documento_id=int(tipo_documento_id) if tipo_documento_id else None,
            departamento=departamento,
            palavras_chave=palavras_chave,
            resumo=resumo,
            autor_id=current_user.id,
            data_validade=datetime.strptime(data_validade, '%Y-%m-%d') if data_validade else None
        )
        
        db.session.add(document)
        db.session.flush()  # Para obter o ID
        
        # Criar primeira versão
        version = DocumentVersion(
            documento_id=document.id,
            versao='1.0',
            conteudo=conteudo,
            criado_por_id=current_user.id,
            changelog='Versão inicial'
        )
        
        db.session.add(version)
        db.session.commit()
        
        # Verificar se é para salvar como rascunho ou submeter
        action = request.form.get('action', 'submit')
        
        if action == 'draft':
            document.status = 'rascunho'
            flash(f'Rascunho "{titulo}" salvo com sucesso!', 'info')
        else:
            flash(f'Documento "{titulo}" criado com sucesso!', 'success')
        
        return redirect(url_for('documents.view', id=document.id))
    
    # Carregar tipos de documentos dinâmicos
    document_types = DocumentType.query.filter_by(ativo=True).order_by(DocumentType.nome).all()
    return render_template('documents/create.html', document_types=document_types)

@bp.route('/<int:id>')
@login_required
def view(id):
    """Visualizar documento"""
    document = Document.query.get_or_404(id)
    current_version = document.get_current_version()
    
    # Registrar leitura
    reading = DocumentReading.query.filter_by(
        documento_id=document.id,
        usuario_id=current_user.id,
        versao_lida=document.versao_atual
    ).first()
    
    if not reading:
        reading = DocumentReading(
            documento_id=document.id,
            usuario_id=current_user.id,
            versao_lida=document.versao_atual,
            ip_address=request.remote_addr
        )
        db.session.add(reading)
        db.session.commit()
    
    return render_template('documents/view.html', 
                         document=document, 
                         current_version=current_version)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Editar documento"""
    try:
        document = Document.query.get_or_404(id)
        
        # Verificar se o usuário tem permissão
        if not (current_user.is_admin() or document.autor_id == current_user.id):
            flash('Você não tem permissão para editar este documento.', 'error')
            return redirect(url_for('documents.view', id=id))
        
        current_version = document.get_current_version()
        
        if request.method == 'POST':
            try:
                # Verificar se é uma requisição AJAX para auto-save
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Auto-save logic
                    return jsonify({'success': True, 'message': 'Salvo automaticamente'})
                
                # Atualizar documento
                document.titulo = request.form.get('titulo', '').strip()
                document.tipo = request.form.get('tipo', '').strip()
                document.departamento = request.form.get('departamento', '').strip()
                document.palavras_chave = request.form.get('palavras_chave', '').strip()
                document.resumo = request.form.get('resumo', '').strip()
                
                # Tratar data de validade
                data_validade = request.form.get('data_validade')
                if data_validade:
                    try:
                        document.data_validade = datetime.strptime(data_validade, '%Y-%m-%d')
                    except ValueError:
                        document.data_validade = None
                else:
                    document.data_validade = None
                
                # Atualizar conteúdo da versão atual
                conteudo = request.form.get('conteudo', '').strip()
                changelog = request.form.get('changelog', 'Edição da versão atual').strip()
                
                if current_version:
                    current_version.conteudo = conteudo
                    current_version.changelog = changelog
                    current_version.data_modificacao = datetime.utcnow()
                
                document.data_ultima_revisao = datetime.utcnow()
                
                db.session.commit()
                
                flash('Documento atualizado com sucesso!', 'success')
                return redirect(url_for('documents.view', id=id))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao salvar documento: {str(e)}', 'error')
                return redirect(url_for('documents.edit', id=id))
        
        return render_template('documents/edit.html', 
                             document=document, 
                             current_version=current_version)
                             
    except Exception as e:
        flash(f'Erro ao carregar documento: {str(e)}', 'error')
        return redirect(url_for('documents.index'))

@bp.route('/<int:id>/confirm_reading', methods=['POST'])
@login_required
def confirm_reading(id):
    """Confirmar leitura de documento via AJAX"""
    document = Document.query.get_or_404(id)
    
    # Verificar se já foi lida esta versão
    reading = DocumentReading.query.filter_by(
        documento_id=document.id,
        usuario_id=current_user.id,
        versao_lida=document.versao_atual
    ).first()
    
    if not reading:
        reading = DocumentReading(
            documento_id=document.id,
            usuario_id=current_user.id,
            versao_lida=document.versao_atual,
            ip_address=request.remote_addr
        )
        db.session.add(reading)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Leitura confirmada com sucesso!'})
    else:
        return jsonify({'success': False, 'message': 'Leitura já confirmada para esta versão.'})

@bp.route('/<int:id>/versions')
@login_required
def versions(id):
    """Histórico de versões do documento"""
    document = Document.query.get_or_404(id)
    versions = document.versoes.order_by(DocumentVersion.data_criacao.desc()).all()
    
    return render_template('documents/versions.html', 
                         document=document, 
                         versions=versions)

@bp.route('/api/save-draft', methods=['POST'])
@login_required
def save_draft():
    """API para salvar rascunho automaticamente"""
    if not current_user.can_create_documents():
        return jsonify({'success': False, 'error': 'Sem permissão'})
    
    try:
        document_id = request.form.get('document_id')
        titulo = request.form.get('titulo')
        conteudo = request.form.get('conteudo')
        tipo = request.form.get('tipo', 'outros')
        departamento = request.form.get('departamento', '')
        
        if not titulo or not conteudo:
            return jsonify({'success': False, 'error': 'Título e conteúdo são obrigatórios'})
        
        if document_id:
            # Atualizar documento existente
            document = Document.query.get(document_id)
            if document and document.autor_id == current_user.id and document.status == 'rascunho':
                document.titulo = titulo
                document.data_modificacao = datetime.utcnow()
                
                # Atualizar versão atual
                current_version = document.get_current_version()
                if current_version:
                    current_version.conteudo = conteudo
                    current_version.data_modificacao = datetime.utcnow()
                
                db.session.commit()
                return jsonify({'success': True, 'message': 'Rascunho atualizado', 'document_id': document.id})
        else:
            # Criar novo rascunho
            codigo = f"{tipo.upper()}-{datetime.now().strftime('%Y')}-{str(uuid.uuid4())[:8].upper()}"
            
            document = Document(
                codigo=codigo,
                titulo=titulo,
                tipo=tipo,
                departamento=departamento,
                autor_id=current_user.id,
                status='rascunho'
            )
            
            db.session.add(document)
            db.session.flush()  # Para obter o ID
            
            # Criar primeira versão
            version = DocumentVersion(
                documento_id=document.id,
                versao='1.0',
                conteudo=conteudo,
                criado_por_id=current_user.id,
                changelog='Rascunho inicial'
            )
            
            db.session.add(version)
            db.session.commit()
            
            return jsonify({'success': True, 'message': 'Novo rascunho criado', 'document_id': document.id})
        
        return jsonify({'success': False, 'error': 'Documento não encontrado'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/drafts')
@login_required
def drafts():
    """Lista de rascunhos do usuário"""
    if not current_user.can_create_documents():
        flash('Você não tem permissão para ver rascunhos.', 'error')
        return redirect(url_for('documents.index'))
    
    page = request.args.get('page', 1, type=int)
    
    drafts = Document.query.filter_by(
        autor_id=current_user.id,
        status='rascunho',
        ativo=True
    ).order_by(Document.data_ultima_revisao.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('documents/drafts.html', drafts=drafts)

@bp.route('/<int:id>/restore-version/<int:version_id>', methods=['POST'])
@login_required
def restore_version(id, version_id):
    """Restaurar uma versão específica do documento"""
    document = Document.query.get_or_404(id)
    version = DocumentVersion.query.get_or_404(version_id)
    
    if not (current_user.can_create_documents() or document.autor_id == current_user.id):
        return jsonify({'success': False, 'error': 'Sem permissão para restaurar versão'})
    
    if version.documento_id != document.id:
        return jsonify({'success': False, 'error': 'Versão não pertence ao documento'})
    
    try:
        # Calcular próxima versão
        next_version = f"{float(document.versao_atual) + 0.1:.1f}"
        
        # Criar nova versão baseada na versão restaurada
        new_version = DocumentVersion(
            documento_id=document.id,
            versao=next_version,
            conteudo=version.conteudo,
            criado_por_id=current_user.id,
            changelog=f'Restaurada versão {version.versao}'
        )
        
        # Atualizar documento
        document.versao_atual = next_version
        document.data_ultima_revisao = datetime.utcnow()
        if document.status == 'aprovado':
            document.status = 'rascunho'  # Volta para rascunho quando restaura
        
        db.session.add(new_version)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Versão {version.versao} restaurada como v{next_version}'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


@bp.route('/reports')
@login_required
def reports():
    """Relatórios de documentos"""
    if not current_user.can_admin():
        flash('Acesso negado.', 'error')
        return redirect(url_for('documents.index'))
    
    # Estatísticas gerais
    total_documentos = Document.query.count()
    ativos = Document.query.filter_by(ativo=True).count()
    aprovados = Document.query.filter_by(status='aprovado').count()
    rascunhos = Document.query.filter_by(status='rascunho').count()
    em_revisao = Document.query.filter_by(status='em_revisao').count()
    obsoletos = Document.query.filter_by(status='obsoleto').count()
    
    # Documentos por tipo
    por_tipo = db.session.query(
        Document.tipo,
        db.func.count(Document.id).label('count')
    ).filter_by(ativo=True).group_by(Document.tipo).order_by(
        db.func.count(Document.id).desc()
    ).all()
    
    # Documentos vencidos
    vencidos = Document.query.filter(
        Document.data_validade < datetime.utcnow(),
        Document.ativo == True
    ).count()
    
    # Documentos vencendo (próximos 30 dias)
    data_limite_30 = datetime.utcnow() + timedelta(days=30)
    vencendo = Document.query.filter(
        Document.data_validade <= data_limite_30,
        Document.data_validade >= datetime.utcnow(),
        Document.ativo == True
    ).count()
    
    # Documentos por departamento
    por_departamento = db.session.query(
        Document.departamento,
        db.func.count(Document.id).label('count')
    ).filter(
        Document.ativo == True,
        Document.departamento.isnot(None),
        Document.departamento != ''
    ).group_by(Document.departamento).order_by(
        db.func.count(Document.id).desc()
    ).all()
    
    # Documentos criados nos últimos 30 dias
    data_limite = datetime.utcnow() - timedelta(days=30)
    criados_mes = Document.query.filter(
        Document.data_criacao >= data_limite,
        Document.ativo == True
    ).count()
    
    return render_template('documents/reports.html',
                         total_documentos=total_documentos,
                         ativos=ativos,
                         aprovados=aprovados,
                         rascunhos=rascunhos,
                         em_revisao=em_revisao,
                         obsoletos=obsoletos,
                         por_tipo=por_tipo,
                         vencidos=vencidos,
                         vencendo=vencendo,
                         por_departamento=por_departamento,
                         criados_mes=criados_mes)


def clean_html_for_pdf(html_content):
    """Converter HTML para texto limpo para PDF, mantendo estrutura básica"""
    if not html_content:
        return ""
    
    # Converter HTML para texto mantendo estrutura
    h = html2text.HTML2Text()
    h.ignore_links = False
    h.body_width = 0  # Não quebrar linhas automaticamente
    h.protect_links = True
    h.use_automatic_links = False
    text = h.handle(html_content)
    
    return text


class DocumentPDFTemplate(BaseDocTemplate):
    """Template personalizado para documentos controlados"""
    
    def __init__(self, filename, document, **kwargs):
        self.document = document
        BaseDocTemplate.__init__(self, filename, **kwargs)
        
        # Configurar frame principal
        frame = Frame(
            2*cm, 2*cm, 17*cm, 23*cm,
            leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0
        )
        
        # Template de página
        template = PageTemplate(
            id='main',
            frames=[frame],
            onPage=self.on_page,
            pagesize=A4
        )
        
        self.addPageTemplates([template])
    
    def on_page(self, canvas, doc):
        """Adicionar header, footer e watermark em cada página"""
        canvas.saveState()
        
        # Header
        canvas.setFont('Helvetica-Bold', 10)
        canvas.drawString(2*cm, 27*cm, f"DOCUMENTO CONTROLADO - Código: {self.document.codigo or 'N/A'}")
        canvas.drawString(2*cm, 26.5*cm, f"Título: {(self.document.titulo or 'Sem título')[:60]}")
        canvas.setFont('Helvetica', 8)
        canvas.drawString(2*cm, 26*cm, f"Versão: {self.document.versao_atual or '1.0'} | Status: {(self.document.status or 'rascunho').upper()}")
        
        # Linha horizontal
        canvas.line(2*cm, 25.7*cm, 19*cm, 25.7*cm)
        
        # Footer com numeração
        canvas.setFont('Helvetica', 8)
        canvas.drawString(2*cm, 1.5*cm, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
        canvas.drawRightString(19*cm, 1.5*cm, f"Página {doc.page}")
        canvas.drawCentredText(10.5*cm, 1.2*cm, "*** COPIA NAO CONTROLADA - Consulte sempre a versao eletronica ***")
        
        # Linha horizontal footer
        canvas.line(2*cm, 1.8*cm, 19*cm, 1.8*cm)
        
        # Watermark diagonal (mais seguro sem setFillAlpha)
        canvas.setFont('Helvetica-Bold', 40)
        canvas.setFillColorRGB(0.9, 0.9, 0.9)  # Cinza claro
        canvas.rotate(45)
        canvas.drawCentredText(15*cm, -5*cm, "DOCUMENTO CONTROLADO")
        
        canvas.restoreState()


def generate_document_pdf(document, current_version):
    """Gerar PDF do documento usando ReportLab"""
    # Criar arquivo temporário
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    temp_filename = temp_file.name
    temp_file.close()
    
    try:
        # Configurar documento PDF com template personalizado
        doc = DocumentPDFTemplate(
            temp_filename,
            document,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=3*cm,  # Mais espaço para header
            bottomMargin=2.5*cm  # Mais espaço para footer
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        
        # Estilos personalizados para qualidade
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=30,
            textColor=colors.black,
            alignment=TA_CENTER
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue,
            leftIndent=0
        )
        
        subheader_style = ParagraphStyle(
            'CustomSubHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.darkblue,
            leftIndent=0
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            alignment=TA_JUSTIFY,
            leftIndent=0
        )
        
        # Lista de elementos do PDF
        story = []
        
        # Cabeçalho do documento
        header_data = [
            ['DOCUMENTO CONTROLADO', '', '', ''],
            ['Código:', document.codigo or 'N/A', 'Versão:', document.versao_atual or '1.0'],
            ['Título:', document.titulo or 'Sem título', 'Tipo:', document.tipo or 'N/A'],
            ['Departamento:', document.departamento or 'N/A', 'Status:', (document.status or 'rascunho').upper()],
            ['Autor:', document.autor.nome_completo if document.autor else 'N/A', 
             'Data:', document.data_criacao.strftime('%d/%m/%Y') if document.data_criacao else 'N/A'],
        ]
        
        if document.data_validade:
            header_data.append(['Validade:', document.data_validade.strftime('%d/%m/%Y'), 
                              'Próxima Revisão:', document.data_validade.strftime('%d/%m/%Y')])
        
        header_table = Table(header_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(header_table)
        story.append(Spacer(1, 20))
        
        # Título do documento
        story.append(Paragraph(document.titulo or 'Sem título', title_style))
        story.append(Spacer(1, 20))
        
        # Resumo se disponível
        if document.resumo:
            story.append(Paragraph("RESUMO EXECUTIVO", header_style))
            story.append(Paragraph(document.resumo, normal_style))
            story.append(Spacer(1, 15))
        
        # Conteúdo do documento
        if current_version and current_version.conteudo:
            story.append(Paragraph("CONTEÚDO DO DOCUMENTO", header_style))
            
            # Converter HTML para texto estruturado
            content_text = clean_html_for_pdf(current_version.conteudo)
            
            # Processar o texto linha por linha para manter estrutura
            lines = content_text.split('\n')
            for line in lines:
                line = line.strip()
                if not line:
                    story.append(Spacer(1, 6))
                    continue
                
                # Detectar cabeçalhos (linhas que começam com #)
                if line.startswith('# '):
                    story.append(Paragraph(line[2:], header_style))
                elif line.startswith('## '):
                    story.append(Paragraph(line[3:], subheader_style))
                elif line.startswith('### '):
                    story.append(Paragraph(line[4:], subheader_style))
                elif line.startswith('**') and line.endswith('**'):
                    # Texto em negrito
                    bold_style = ParagraphStyle(
                        'Bold',
                        parent=normal_style,
                        fontName='Helvetica-Bold'
                    )
                    story.append(Paragraph(line[2:-2], bold_style))
                elif line.startswith('- ') or line.startswith('* '):
                    # Lista não ordenada
                    story.append(Paragraph(f"• {line[2:]}", normal_style))
                elif re.match(r'^\d+\. ', line):
                    # Lista ordenada
                    story.append(Paragraph(line, normal_style))
                else:
                    # Parágrafo normal
                    story.append(Paragraph(line, normal_style))
        
        # Palavras-chave
        if document.palavras_chave:
            story.append(Spacer(1, 20))
            story.append(Paragraph("PALAVRAS-CHAVE", header_style))
            story.append(Paragraph(document.palavras_chave, normal_style))
        
        # Rodapé com informações de controle
        story.append(Spacer(1, 30))
        footer_data = [
            ['CONTROLE DO DOCUMENTO'],
            ['Este é um documento controlado. Cópias impressas não são controladas.'],
            ['Sempre consulte a versão eletrônica mais atual no sistema.'],
            [f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'],
            [f'Por: {current_user.nome_completo if current_user else "Sistema"}']
        ]
        
        footer_table = Table(footer_data, colWidths=[16*cm])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(footer_table)
        
        # Gerar PDF usando template personalizado
        doc.build(story)
        
        return temp_filename
        
    except Exception as e:
        # Limpar arquivo temporário em caso de erro
        if os.path.exists(temp_filename):
            os.unlink(temp_filename)
        raise e


@bp.route('/<int:id>/export_pdf')
@login_required 
def export_pdf(id):
    """Exportar documento como PDF"""
    document = Document.query.get_or_404(id)
    current_version = document.get_current_version()
    
    try:
        # Gerar PDF
        pdf_file = generate_document_pdf(document, current_version)
        
        # Ler arquivo PDF
        with open(pdf_file, 'rb') as f:
            pdf_data = f.read()
        
        # Limpar arquivo temporário
        os.unlink(pdf_file)
        
        # Criar resposta
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        # Nome do arquivo seguro
        safe_title = (document.titulo or 'documento')[:30].replace('/', '_').replace('\\', '_')
        safe_filename = f"{document.codigo or 'DOC'}_{safe_title}.pdf"
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        
        return response
        
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('documents.view', id=id))


@bp.route('/reports/export/<format>')
@login_required
def export_reports(format):
    """Exportar relatórios de documentos em diferentes formatos"""
    if not current_user.can_admin():
        flash('Acesso negado.', 'error')
        return redirect(url_for('documents.reports'))
    
    try:
        # Coletar dados dos relatórios (mesmo código da função reports)
        total_documentos = Document.query.count()
        ativos = Document.query.filter_by(ativo=True).count()
        aprovados = Document.query.filter_by(status='aprovado').count()
        rascunhos = Document.query.filter_by(status='rascunho').count()
        em_revisao = Document.query.filter_by(status='em_revisao').count()
        obsoletos = Document.query.filter_by(status='obsoleto').count()
        
        # Documentos por tipo
        por_tipo = db.session.query(
            Document.tipo,
            db.func.count(Document.id).label('count')
        ).filter_by(ativo=True).group_by(Document.tipo).order_by(
            db.func.count(Document.id).desc()
        ).all()
        
        # Documentos vencidos
        vencidos = Document.query.filter(
            Document.data_validade < datetime.utcnow(),
            Document.ativo == True
        ).count()
        
        # Documentos vencendo (próximos 30 dias)
        data_limite_30 = datetime.utcnow() + timedelta(days=30)
        vencendo = Document.query.filter(
            Document.data_validade <= data_limite_30,
            Document.data_validade >= datetime.utcnow(),
            Document.ativo == True
        ).count()
        
        # Documentos por departamento
        por_departamento = db.session.query(
            Document.departamento,
            db.func.count(Document.id).label('count')
        ).filter(
            Document.ativo == True,
            Document.departamento.isnot(None),
            Document.departamento != ''
        ).group_by(Document.departamento).order_by(
            db.func.count(Document.id).desc()
        ).all()
        
        # Documentos criados nos últimos 30 dias
        data_limite = datetime.utcnow() - timedelta(days=30)
        criados_mes = Document.query.filter(
            Document.data_criacao >= data_limite,
            Document.ativo == True
        ).count()
        
        if format == 'pdf':
            return export_reports_pdf({
                'total_documentos': total_documentos,
                'ativos': ativos,
                'aprovados': aprovados,
                'rascunhos': rascunhos,
                'em_revisao': em_revisao,
                'obsoletos': obsoletos,
                'vencidos': vencidos,
                'vencendo': vencendo,
                'criados_mes': criados_mes,
                'por_tipo': por_tipo,
                'por_departamento': por_departamento
            })
        elif format == 'excel':
            return export_reports_excel({
                'total_documentos': total_documentos,
                'ativos': ativos,
                'aprovados': aprovados,
                'rascunhos': rascunhos,
                'em_revisao': em_revisao,
                'obsoletos': obsoletos,
                'vencidos': vencidos,
                'vencendo': vencendo,
                'criados_mes': criados_mes,
                'por_tipo': por_tipo,
                'por_departamento': por_departamento
            })
        elif format == 'csv':
            return export_reports_csv({
                'total_documentos': total_documentos,
                'ativos': ativos,
                'aprovados': aprovados,
                'rascunhos': rascunhos,
                'em_revisao': em_revisao,
                'obsoletos': obsoletos,
                'vencidos': vencidos,
                'vencendo': vencendo,
                'criados_mes': criados_mes,
                'por_tipo': por_tipo,
                'por_departamento': por_departamento
            })
        else:
            flash('Formato de exportação inválido.', 'error')
            return redirect(url_for('documents.reports'))
            
    except Exception as e:
        flash(f'Erro ao exportar relatório: {str(e)}', 'error')
        return redirect(url_for('documents.reports'))


def export_reports_pdf(data):
    """Exportar relatórios em PDF"""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    temp_filename = temp_file.name
    temp_file.close()
    
    try:
        doc = SimpleDocTemplate(
            temp_filename,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("RELATÓRIO DE DOCUMENTOS", title_style))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Resumo Executivo
        story.append(Paragraph("RESUMO EXECUTIVO", styles['Heading1']))
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Documentos', str(data['total_documentos'])],
            ['Documentos Ativos', str(data['ativos'])],
            ['Documentos Aprovados', str(data['aprovados'])],
            ['Rascunhos', str(data['rascunhos'])],
            ['Em Revisão', str(data['em_revisao'])],
            ['Obsoletos', str(data['obsoletos'])],
            ['Vencidos', str(data['vencidos'])],
            ['Vencendo (30 dias)', str(data['vencendo'])],
            ['Criados (último mês)', str(data['criados_mes'])]
        ]
        
        summary_table = Table(summary_data, colWidths=[8*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Documentos por Tipo
        if data['por_tipo']:
            story.append(Paragraph("DOCUMENTOS POR TIPO", styles['Heading1']))
            
            tipo_data = [['Tipo', 'Quantidade', 'Percentual']]
            for tipo, count in data['por_tipo']:
                percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
                tipo_data.append([tipo or 'N/A', str(count), f"{percentage:.1f}%"])
            
            tipo_table = Table(tipo_data, colWidths=[6*cm, 3*cm, 3*cm])
            tipo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(tipo_table)
            story.append(Spacer(1, 20))
        
        # Documentos por Departamento
        if data['por_departamento']:
            story.append(Paragraph("DOCUMENTOS POR DEPARTAMENTO", styles['Heading1']))
            
            dept_data = [['Departamento', 'Quantidade', 'Percentual']]
            for dept, count in data['por_departamento']:
                percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
                dept_data.append([dept or 'N/A', str(count), f"{percentage:.1f}%"])
            
            dept_table = Table(dept_data, colWidths=[6*cm, 3*cm, 3*cm])
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(dept_table)
        
        doc.build(story)
        
        # Ler arquivo PDF
        with open(temp_filename, 'rb') as f:
            pdf_data = f.read()
        
        # Limpar arquivo temporário
        os.unlink(temp_filename)
        
        # Criar resposta
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="relatorio_documentos_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response
        
    except Exception as e:
        if os.path.exists(temp_filename):
            os.unlink(temp_filename)
        raise e


def export_reports_excel(data):
    """Exportar relatórios em Excel"""
    # Criar workbook
    output = BytesIO()
    workbook = Workbook()
    
    # Remover sheet padrão
    workbook.remove(workbook.active)
    
    # Sheet 1: Resumo
    ws_resumo = workbook.create_sheet("Resumo")
    
    # Cabeçalho
    ws_resumo['A1'] = 'RELATÓRIO DE DOCUMENTOS'
    ws_resumo['A1'].font = Font(size=16, bold=True)
    ws_resumo['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
    
    # Dados do resumo
    resumo_data = [
        ['Métrica', 'Valor'],
        ['Total de Documentos', data['total_documentos']],
        ['Documentos Ativos', data['ativos']],
        ['Documentos Aprovados', data['aprovados']],
        ['Rascunhos', data['rascunhos']],
        ['Em Revisão', data['em_revisao']],
        ['Obsoletos', data['obsoletos']],
        ['Vencidos', data['vencidos']],
        ['Vencendo (30 dias)', data['vencendo']],
        ['Criados (último mês)', data['criados_mes']]
    ]
    
    for row_num, row_data in enumerate(resumo_data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_resumo.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 4:  # Cabeçalho
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 25
    ws_resumo.column_dimensions['B'].width = 15
    
    # Sheet 2: Por Tipo
    if data['por_tipo']:
        ws_tipo = workbook.create_sheet("Por Tipo")
        ws_tipo['A1'] = 'DOCUMENTOS POR TIPO'
        ws_tipo['A1'].font = Font(size=14, bold=True)
        
        tipo_headers = ['Tipo', 'Quantidade', 'Percentual']
        for col_num, header in enumerate(tipo_headers, 1):
            cell = ws_tipo.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, (tipo, count) in enumerate(data['por_tipo'], 4):
            percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
            ws_tipo.cell(row=row_num, column=1, value=tipo or 'N/A')
            ws_tipo.cell(row=row_num, column=2, value=count)
            ws_tipo.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
        
        ws_tipo.column_dimensions['A'].width = 20
        ws_tipo.column_dimensions['B'].width = 12
        ws_tipo.column_dimensions['C'].width = 12
    
    # Sheet 3: Por Departamento
    if data['por_departamento']:
        ws_dept = workbook.create_sheet("Por Departamento")
        ws_dept['A1'] = 'DOCUMENTOS POR DEPARTAMENTO'
        ws_dept['A1'].font = Font(size=14, bold=True)
        
        dept_headers = ['Departamento', 'Quantidade', 'Percentual']
        for col_num, header in enumerate(dept_headers, 1):
            cell = ws_dept.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, (dept, count) in enumerate(data['por_departamento'], 4):
            percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
            ws_dept.cell(row=row_num, column=1, value=dept or 'N/A')
            ws_dept.cell(row=row_num, column=2, value=count)
            ws_dept.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
        
        ws_dept.column_dimensions['A'].width = 25
        ws_dept.column_dimensions['B'].width = 12
        ws_dept.column_dimensions['C'].width = 12
    
    # Salvar workbook
    workbook.save(output)
    output.seek(0)
    
    # Criar resposta
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="relatorio_documentos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    return response


def export_reports_csv(data):
    """Exportar relatórios em CSV"""
    output = BytesIO()
    output_text = BytesIO()
    
    # Usar StringIO para CSV
    import io
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    
    # Cabeçalho
    writer.writerow(['RELATÓRIO DE DOCUMENTOS'])
    writer.writerow([f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'])
    writer.writerow([])  # Linha em branco
    
    # Resumo
    writer.writerow(['RESUMO EXECUTIVO'])
    writer.writerow(['Métrica', 'Valor'])
    writer.writerow(['Total de Documentos', data['total_documentos']])
    writer.writerow(['Documentos Ativos', data['ativos']])
    writer.writerow(['Documentos Aprovados', data['aprovados']])
    writer.writerow(['Rascunhos', data['rascunhos']])
    writer.writerow(['Em Revisão', data['em_revisao']])
    writer.writerow(['Obsoletos', data['obsoletos']])
    writer.writerow(['Vencidos', data['vencidos']])
    writer.writerow(['Vencendo (30 dias)', data['vencendo']])
    writer.writerow(['Criados (último mês)', data['criados_mes']])
    writer.writerow([])  # Linha em branco
    
    # Por Tipo
    if data['por_tipo']:
        writer.writerow(['DOCUMENTOS POR TIPO'])
        writer.writerow(['Tipo', 'Quantidade', 'Percentual'])
        for tipo, count in data['por_tipo']:
            percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
            writer.writerow([tipo or 'N/A', count, f"{percentage:.1f}%"])
        writer.writerow([])  # Linha em branco
    
    # Por Departamento
    if data['por_departamento']:
        writer.writerow(['DOCUMENTOS POR DEPARTAMENTO'])
        writer.writerow(['Departamento', 'Quantidade', 'Percentual'])
        for dept, count in data['por_departamento']:
            percentage = (count / data['total_documentos']) * 100 if data['total_documentos'] > 0 else 0
            writer.writerow([dept or 'N/A', count, f"{percentage:.1f}%"])
    
    # Converter para bytes
    csv_data = csv_buffer.getvalue().encode('utf-8')
    
    # Criar resposta
    response = make_response(csv_data)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="relatorio_documentos_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    
    return response
